#-*- coding:utf-8 –*-
#!/usr/bin/env python
#-------------------------------------------------------------------------------
# Purpose: 下载Peer数据库的全部力位移关系文件
# Author: Bill
# Created: 23/5/2016
#-------------------------------------------------------------------------------
from openpyxl import load_workbook
import glob
import matplotlib.gridspec as gridspec
import matplotlib.pyplot as plt
import numpy as np
import os
import re
import string
import subprocess 
import urllib
#-------------------------------------------------------------------------------
PI=3.14159
#b=720
#h=420
#c=30
#As=PI*8**2
#m=2
#n=3
#fy=400
#fc=58.
#ft=0
#size=10
#P=-120000
#s=100
#-------------------------------------------------------------------------------
# 弯矩曲率分析
#-------------------------------------------------------------------------------
def dbmax(idd):
	filelist=os.listdir('rectcol')
	filename=0
	for a in filelist:
		if a.startswith('%03d' %(idd)):filename=a
	if filename==0:return 0
	VD=np.loadtxt('rectcol\\'+filename,skiprows=2)
	ans = np.max(VD[:,1])
	ind=np.min(np.where(VD[:,1]==ans))
	return [VD[ind,0],ans*1000]
def dmax(idd):
	filelist=os.listdir('rectcol')
	filename=0
	for a in filelist:
		if a.startswith('%03d' %(idd)):filename=a
	if filename==0:return 0
	VD=np.loadtxt('rectcol\\'+filename,skiprows=2)
	ans = np.max(np.abs(VD[:,0]))
	return ans
def MyKy(b,h,c,As,m,n,fy,fc,ft,size,P):
	fid=open('para.tcl','w+')
	fid.write("set b %f;\n"%(b))
	fid.write("set h %f;\n"%(h))
	fid.write("set c %f;\n"%(c))
	fid.write("set As %f;\n"%(As))
	fid.write("set m %d;\n"%(m))
	fid.write("set n %d;\n"%(n))
	fid.write("set fy %f;\n"%(fy))
	fid.write("set fc %f;\n"%(fc))
	fid.write("set ft %f;\n"%(ft))
	fid.write("set size %f;\n"%(size))
	fid.write("set Pload %f;"%(P))
	fid.close()
	p=subprocess.Popen('12.cmd',shell=False)
	p.wait()
	ANS=np.loadtxt('INF\\res.out')
	mk=np.loadtxt('INF\\Mk.out')
	M3=np.max(mk[:,0])
	K3=np.min(mk[np.where(mk[:,0]==M3),1])
	return np.hstack((ANS,[M3,K3]))
def vaci(b,h,c,At,nt,fyt,fc,s,P):
	Vc=(1+P/(14*b*h))*(np.sqrt(fc)/6)*b*(h-c)
	Vs=At*nt*fyt*(h-c)/s
	return Vc+Vs
def pri(b,h,c,At,nt,fyt,fc,s,P,L):
	Vc=0.29*np.sqrt(fc)*0.8*b*h
	Vs=At*nt*fyt*(h-c*2)/s*1.732
	Vp=(h-0.15*h)/2/L*P
	return Vc+Vs+Vp
def pri2(b,h,c,At,nt,fyt,fc,s,P,L):
	Vc=0.1*np.sqrt(fc)*0.8*b*h
	Vs=At*nt*fyt*(h-c*2)/s*1.732
	Vp=(h-0.15*h)/2/L*P
	return Vc+Vs+Vp
def sezen(b,h,c,At,nt,fyt,fc,s,P,L):
	mmin=0.0393701
	mpsi=145.0377439
	nlb=0.2248089
	b=b*mmin
	h=h*mmin
	c=c*mmin
	At=At*mmin*mmin
	fyt=fyt*mpsi
	fc=fc*mpsi
	s=s*mmin
	P=P*nlb
	L=L*mmin
	Vc=6*np.sqrt(fc)/(L/(h-c))*np.sqrt(1+P/(6*np.sqrt(fc)*b*h))*0.8*b*h
	Vs=At*nt*fyt*(h-c)/s
	return (Vc+Vs)/nlb
def main1():#进行数据库计算
	c=30
	size=10
	ft=0	
	wb = load_workbook(filename='TestPara.xlsx')
	st=wb['Sheet2']
	st['Q1']="Vy"
	st['R1']="Vmax"
	for aa in range(301):
		num=st['A'+str(aa+2)].value
		h=st['B'+str(aa+2)].value
		b=st['C'+str(aa+2)].value
		As=st['D'+str(aa+2)].value**2*PI*0.25
		ns=st['E'+str(aa+2)].value
		m=st['F'+str(aa+2)].value
		n=(ns-4-2*m)/2
		fy=st['M'+str(aa+2)].value
		fc=st['L'+str(aa+2)].value
		#P=st['J'+str(aa+2)].value
		P=0
		ch=st['P'+str(aa+2)].value
		L=st['K'+str(aa+2)].value
		FM=st['O'+str(aa+2)].value
		if ch!=0 and FM==1 : 
			res=MyKy(b,h,c,As,m,n,fy,fc,ft,size,P)
			st['Q'+str(aa+2)]=res[0]/L
			st['R'+str(aa+2)]=dbmax(num)[1]
		wb.save('TestPara.xlsx')	
#-------------------------------------------------------------------------------
# 抗剪计算(试验)
#-------------------------------------------------------------------------------
def shearCal():
	c=30
	size=10
	ft=0
	fyt=235
	fy=400
	fc=58.
	s=100
	wb = load_workbook(filename='TestPara.xlsx')
	sheet=wb['Sheet1']
	sheet['L1']="Vmax"
	sheet['M1']="Vaci"
	sheet['N1']="Vpri"
	sheet['O1']="Vsen"
	for x in range(18):
		h=sheet['B'+str(x+2)].value
		b=sheet['C'+str(x+2)].value
		ds=sheet['D'+str(x+2)].value
		As=ds**2*3.14159*0.25
		ns=sheet['E'+str(x+2)].value
		m=sheet['F'+str(x+2)].value
		n=(ns-4-2*m)/2
		At=sheet['G'+str(x+2)].value**2*3.14159*0.25
		nt=sheet['H'+str(x+2)].value
		s=sheet['I'+str(x+2)].value
		P=sheet['J'+str(x+2)].value*1000
		L=sheet['K'+str(x+2)].value
		mk=MyKy(b,h,c,As,m,n,fy,fc,ft,size,P)
		sheet['L'+str(x+2)]=mk[0]/L*2.0#这是我得出来的经验系数
		sheet['M'+str(x+2)]=vaci(b,h,c,At,nt,fyt,fc,s,P)
		sheet['N'+str(x+2)]=pri(b,h,c,At,nt,fyt,fc,s,P,L)
		sheet['O'+str(x+2)]=sezen(b,h,c,At,nt,fyt,fc,s,P,L)
	wb.save('TestPara.xlsx')
#-------------------------------------------------------------------------------
# 抗剪能力验证(数据库)
#-------------------------------------------------------------------------------
def shear_db():
	size=10
	ft=0	
	wb = load_workbook(filename='TestPara.xlsx')
	st=wb['Sheet2']
	st['R1']="Vmax"
	st['S1']="Vaci"
	st['T1']="Vpri"
	st['U1']="Vsen"
	st['V1']="PredVmax"
	for aa in range(301):
		num=st['A'+str(aa+2)].value
		h=st['B'+str(aa+2)].value
		b=st['C'+str(aa+2)].value
		As=st['D'+str(aa+2)].value**2*PI*0.25
		At=st['G'+str(aa+2)].value**2*PI*0.25
		nt=st['H'+str(aa+2)].value
		s=st['I'+str(aa+2)].value
		ns=st['E'+str(aa+2)].value
		m=st['F'+str(aa+2)].value
		n=(ns-4-2*m)/2
		fy=st['M'+str(aa+2)].value
		fyt=st['N'+str(aa+2)].value
		fc=st['L'+str(aa+2)].value
		P=st['J'+str(aa+2)].value
		#P=0
		ch=st['P'+str(aa+2)].value
		L=st['K'+str(aa+2)].value
		FM=st['O'+str(aa+2)].value
		c=st['Q'+str(aa+2)].value
		if ch!=0 and FM==1 : 
			res=MyKy(b,h,c,As,m,n,fy,fc,ft,size,P)
			st['V'+str(aa+2)]=res[0]/L
			print num
			st['R'+str(aa+2)]=dbmax(num)[1]
			st['S'+str(aa+2)]=vaci(b,h,c,At,nt,fyt,fc,s,P)
			st['T'+str(aa+2)]=pri(b,h,c,At,nt,fyt,fc,s,P,L)
			st['U'+str(aa+2)]=sezen(b,h,c,At,nt,fyt,fc,s,P,L)
		wb.save('TestPara.xlsx')
#-------------------------------------------------------------------------------
# 屈服位移估计(Sezen)
#-------------------------------------------------------------------------------
def dyield_Se():
	size=10
	ft=0	
	wb = load_workbook(filename='TestPara.xlsx')
	st=wb['Sheet3']
	st['V1']="Dflex"
	st['W1']="Dslip"
	st['X1']="Dshear"
	st['Y1']="Dy"
	for aa in range(2):
		num=st['A'+str(aa+2)].value
		h=st['B'+str(aa+2)].value
		b=st['C'+str(aa+2)].value
		ds=st['D'+str(aa+2)].value
		As=ds**2*PI*0.25
		At=st['G'+str(aa+2)].value**2*PI*0.25
		nt=st['H'+str(aa+2)].value
		s=st['I'+str(aa+2)].value
		ns=st['E'+str(aa+2)].value
		m=st['F'+str(aa+2)].value
		n=(ns-4-2*m)/2
		fy=st['M'+str(aa+2)].value
		fyt=st['N'+str(aa+2)].value
		fc=st['L'+str(aa+2)].value
		P=st['J'+str(aa+2)].value
		#P=0
		ch=st['P'+str(aa+2)].value
		L=st['K'+str(aa+2)].value
		FM=st['O'+str(aa+2)].value
		c=st['Q'+str(aa+2)].value
		if True: 
			mk=MyKy(b,h,c,As,m,n,fy,fc,ft,size,P)
			phiy=mk[4]*mk[1]/mk[0]
			Dflex=phiy*L**2/3
			Dslip=L*ds*fy*phiy/8/(1.0*np.sqrt(fc))
			Dshear=1.2*mk[4]/(b*h*0.4*5000*np.sqrt(fc))
			st['V'+str(aa+2)]=Dflex
			st['W'+str(aa+2)]=Dslip
			st['X'+str(aa+2)]=Dshear
			st['Y'+str(aa+2)]=Dflex+Dshear+Dslip
		wb.save('TestPara.xlsx')
#-------------------------------------------------------------------------------
# 屈服位移计算（试件）
#-------------------------------------------------------------------------------
def dyield():
	c=30
	size=10
	ft=0
	fyt=235
	fy=400
	fc=58
	s=100
	wb = load_workbook(filename='TestPara.xlsx')
	sheet=wb['Sheet1']
	sheet['P1']="Dflex"
	sheet['Q1']="Dslip"
	sheet['R1']="Dshear"
	sheet['S1']="Dy"
	sheet['T1']="Ky"
	for x in range(18):
		h=sheet['B'+str(x+2)].value
		b=sheet['C'+str(x+2)].value
		ds=sheet['D'+str(x+2)].value
		As=ds**2*3.14159*0.25
		ns=sheet['E'+str(x+2)].value
		m=sheet['F'+str(x+2)].value
		n=(ns-4-2*m)/2
		At=sheet['G'+str(x+2)].value**2*3.14159*0.25
		nt=sheet['H'+str(x+2)].value
		s=sheet['I'+str(x+2)].value
		P=sheet['J'+str(x+2)].value*1000
		L=sheet['K'+str(x+2)].value
		mk=MyKy(b,h,c,As,m,n,fy,fc,ft,size,P)
		phiy=mk[4]*mk[1]/mk[0]
		Dflex=phiy*L**2/3
		Dslip=L*ds*fy*phiy/8/(1.0*np.sqrt(fc))
		Dshear=1.2*mk[4]/(b*h*0.4*5000*np.sqrt(fc))
		sheet['P'+str(x+2)]=Dflex
		sheet['Q'+str(x+2)]=Dslip
		sheet['R'+str(x+2)]=Dshear
		sheet['S'+str(x+2)]=Dflex+Dshear+Dslip
		sheet['T'+str(x+2)]=phiy
	wb.save('TestPara.xlsx')
#-------------------------------------------------------------------------------
# 开裂位移计算（试件）
#-------------------------------------------------------------------------------
def dcr():
	c=30
	size=10
	ft=0
	fyt=235
	fy=400
	fc=58
	s=100
	wb = load_workbook(filename='TestPara.xlsx')
	sheet=wb['Sheet1']
	sheet['U1']="Dcr"
	for x in range(18):
		h=sheet['B'+str(x+2)].value
		b=sheet['C'+str(x+2)].value
		ds=sheet['D'+str(x+2)].value
		As=ds**2*3.14159*0.25
		ns=sheet['E'+str(x+2)].value
		m=sheet['F'+str(x+2)].value
		n=(ns-4-2*m)/2
		At=sheet['G'+str(x+2)].value**2*3.14159*0.25
		nt=sheet['H'+str(x+2)].value
		s=sheet['I'+str(x+2)].value
		P=sheet['J'+str(x+2)].value*1000
		L=sheet['K'+str(x+2)].value
		sigc=np.sqrt(fc)
		Isec=b*h**3/12
		Mcr=(sigc+P/(b*h))*Isec/(h/2)
		Fcr=Mcr/L
		Fci=(0.067+10*(At*nt/s/b))*(1+3*P/(fc*b*h))*b*(h-c)*np.sqrt(fc)
		print Fci
		Econ=5000*np.sqrt(fc)
		Dcr=Fci*L**3/(3*Econ*Isec)
		sheet['U'+str(x+2)]=Dcr
	wb.save('TestPara.xlsx')
#-------------------------------------------------------------------------------
# 单独试件的响应预测(对应Priestely)
#-------------------------------------------------------------------------------
def PreDic(spi='C2',fc=58.0):
	#spi='G-2'
	c=30
	size=10
	ft=np.sqrt(fc)
	fyt=235
	fy=400
	#fc=58
	wb = load_workbook(filename='TestPara.xlsx')
	sheet=wb['Sheet1']
	for x in range(50):
		if sheet['A'+str(x+2)].value==spi:
			h=sheet['B'+str(x+2)].value
			b=sheet['C'+str(x+2)].value
			ds=sheet['D'+str(x+2)].value
			As=ds**2*3.14159*0.25
			ns=sheet['E'+str(x+2)].value
			m=sheet['F'+str(x+2)].value
			n=(ns-4-2*m)/2
			At=sheet['G'+str(x+2)].value**2*3.14159*0.25
			nt=sheet['H'+str(x+2)].value
			s=sheet['I'+str(x+2)].value
			P=sheet['J'+str(x+2)].value*1000
			L=sheet['K'+str(x+2)].value
			mk=MyKy(b,h,c,As,m,n,fy,fc,ft,size,P)
			V=mk[0]/L
			P1=[0.,1.25*V]
			P2=[4.,1.25*V]
			P3=[0.,1.85*V]
			P4=[4.,1.85*V]
			Lb=np.array([P1,P2])
			Gb=np.array([P3,P4])
			Vpri1=pri(b,h,c,At,nt,fyt,fc,s,P,L)
			Vpri2=pri2(b,h,c,At,nt,fyt,fc,s,P,L)
			V1=[0.,Vpri1]
			V2=[2.0,Vpri1]
			V3=[4.0,Vpri2]
			V4=[6.0,Vpri2]
			pp=np.array([V1,V2,V3,V4])
			break
	return (Lb,Gb,pp)#

#-------------------------------------------------------------------------------
# 单独试件的Priestely剪力
#-------------------------------------------------------------------------------
def SheerEQNS(spi='C2',fc=58.0):
	#spi='G-2'
	c=30
	size=10
	ft=np.sqrt(fc)
	fyt=235
	fy=400
	#fc=58
	wb = load_workbook(filename='TestPara.xlsx')
	sheet=wb['Sheet1']
	for x in range(50):
		if sheet['A'+str(x+2)].value==spi:
			h=sheet['B'+str(x+2)].value
			b=sheet['C'+str(x+2)].value
			ds=sheet['D'+str(x+2)].value
			As=ds**2*3.14159*0.25
			ns=sheet['E'+str(x+2)].value
			m=sheet['F'+str(x+2)].value
			n=(ns-4-2*m)/2
			At=sheet['G'+str(x+2)].value**2*3.14159*0.25
			nt=sheet['H'+str(x+2)].value
			s=sheet['I'+str(x+2)].value
			P=sheet['J'+str(x+2)].value*1000
			L=sheet['K'+str(x+2)].value
			Vpri1=pri(b,h,c,At,nt,fyt,fc,s,P,L)
			Vpri2=pri2(b,h,c,At,nt,fyt,fc,s,P,L)
			Vsezen=sezen(b,h,c,At,nt,fyt,fc,s,P,L)
			Vai=vaci(b,h,c,At,nt,fyt,fc,s,P)
			V1=[0.,Vpri1]
			V2=[2.0,Vpri1]
			V3=[4.0,Vpri2]
			V4=[6.0,Vpri2]
			Vpr=np.array([V1,V2,V3,V4])
			Vse=np.array([[0.,Vsezen],[6.0,Vsezen]])
			Vaci=np.array([[0.,Vai],[6.0,Vai]])
			break
	return (Vpr,Vse,Vaci)
#-------------------------------------------------------------------------------
# 3x3图像
#-------------------------------------------------------------------------------
#nameLis=['C-1','C-2','C-3']
#nameLis=['C-2','D-1','D-2','D-3']
#nameLis=['C-2','D-2','G-1','G-2','G-3','G-4']
#nameLis=['C1','C2','C3','E1','E2','G1','G2','D1','D2','D3','F1','F2']
#nn=0
#plt.figure(1,figsize=(20,20))
#for name in nameLis:
#	(L3,L1)=Priestely(name,fc=58.0)
#	(pD,pF)=MyHys(name)
#	fmax=max(pF)
#	nn=nn+1
#	plt.subplot(4,4,nn)
#	plt.plot([0,6],[fmax,fmax],'b',linewidth=2)
#	plt.plot(L3[:,0],L3[:,1],'r')
#	plt.plot(L1[:,0],L1[:,1],'r',alpha=0.5)
#	plt.title('Predict of %s'%(name))
#plt.savefig(u'PyRES\\'+''.join(nameLis)+'.png', dpi=300)
#-------------------------------------------------------------------------------
# 钢筋矩阵
#-------------------------------------------------------------------------------
#dslist=[12,16,18,20,22]
#y=len(dslist)
#slist=[220,150,100,80,70]
#x=len(slist)
#llist=[840]
#h=420
#b=720
#ns=14
#m=2
#n=(ns-4-2*m)/2
#nt=3
#At=0.25*3.14159*36
#P=120
#c=30
#size=10
#fc=58.
#ft=np.sqrt(fc)
#fyt=235
#fy=400
#for L in llist:
#	plt.figure(L,figsize=(y*5,x*5))
#	nn=0
#	for ds in dslist:
#		for s in slist:
#			Vpri1=pri(b,h,c,At,nt,fyt,fc,s,P,L)
#			Vpri2=pri2(b,h,c,At,nt,fyt,fc,s,P,L)
#			Vsezen=sezen(b,h,c,At,nt,fyt,fc,s,P,L)
#			V1=[0.,Vpri1]
#			V2=[2.0,Vpri1]
#			V3=[4.0,Vpri2]
#			V4=[6.0,Vpri2]
#			Vpr=np.array([V1,V2,V3,V4])
#			Vse=np.array([[0.,Vsezen],[6.0,Vsezen]])
#			(pD,pF)=MyHys2(h,b,ds,ns,m,At,nt,s,P,L)
#			fmax=max(pF)
#			nn=nn+1
#			plt.subplot(y,x,nn)
#			plt.plot([0,6],[fmax,fmax],'b',linewidth=2)
#			plt.plot(Vpr[:,0],Vpr[:,1],'r')
#			plt.plot(Vse[:,0],Vse[:,1],'r',alpha=0.5)
#			plt.title('Predict of %s-%s'%(str(ds),str(s)))
#	plt.savefig(u'PyRES\\'+str(L)+'.png', dpi=300)